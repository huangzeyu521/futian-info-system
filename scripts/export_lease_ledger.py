# -*- coding: utf-8 -*-
"""从 materials 目录读取「租赁台账（新）.xlsx」，导出为 UTF-8 JSON。"""
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("请安装: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

MATERIALS = os.path.join(os.path.dirname(__file__), "..", "materials")
OUT = os.path.join(os.path.dirname(__file__), "..", "public", "data", "lease-ledger.json")


def find_xlsx():
    for name in os.listdir(MATERIALS):
        if name.endswith(".xlsx"):
            return os.path.join(MATERIALS, name)
    raise FileNotFoundError("materials 下未找到 .xlsx")


def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(v)
    s = str(v).strip()
    return s


def sheet_to_rows(ws, max_row=5000):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell_str(c) for c in row])
        if len(rows) >= max_row:
            break
    return rows


def parse_baiwang_ruzhu(rows):
    """百旺产业园企业入驻明细：表头约在第2行（索引1），数据从第4行起。"""
    header_idx = 1
    headers = rows[header_idx] if len(rows) > header_idx else []
    # 清洗表头：合并单元格可能只在第一列有标题
    title = rows[0][0] if rows else ""
    records = []
    for i in range(3, len(rows)):
        r = rows[i]
        if not any(r[:8]):
            continue
        # 跳过说明行
        if r[0] and "合计" in r[0]:
            break
        records.append(
            {
                "serialNo": r[0],
                "zone": r[1],
                "buildingRoom": r[2],
                "tenantName": r[3],
                "propertyArea": r[4],
                "allocatedArea": r[5],
                "marketRentArea": r[6],
                "rentUnitPrice": r[7],
                "leaseUnitPrice": r[8],
                "year1RentAfterDiscount": r[9],
                "year2RentAfterDiscount": r[10],
                "year3RentAfterDiscount": r[11],
                "year4RentAfterDiscount": r[12],
                "year5RentAfterDiscount": r[13],
                "startDate": r[14],
                "endDate": r[15],
                "leaseTerm": r[16],
                "increaseRateYear2": r[17],
                "increaseRateOther": r[18],
                "policyBasis": r[19],
                "rights": r[20],
                "remark": r[21],
            }
        )
    return {"title": title, "headers": headers, "records": records}


def parse_baiwang_shangye(rows):
    title = rows[0][0] if rows else ""
    headers = rows[1] if len(rows) > 1 else []
    records = []
    for i in range(2, len(rows)):
        r = rows[i]
        if not r[0] and not r[1]:
            continue
        if r[0] and "合计" in r[0]:
            records.append({"_summaryRow": True, "raw": r})
            continue
        if "绿色" in (r[0] or "") or "未委托" in (r[0] or ""):
            records.append({"_noteRow": True, "raw": r})
            continue
        records.append(
            {
                "serialNo": r[0],
                "address": r[1],
                "rights": r[2],
                "area": r[3],
                "floorOrNo": r[4],
                "status": r[5],
                "progress": r[6],
                "remark": r[7],
            }
        )
    return {"title": title, "headers": headers, "records": records}


def parse_baiwang_zujin(rows):
    title = rows[0][0] if rows else ""
    records = []
    for i in range(1, len(rows)):
        r = rows[i]
        if not any(r[:4]):
            continue
        records.append(
            {
                "label": r[0],
                "value": r[1],
                "extra": r[2:8],
            }
        )
    return {"title": title, "records": records}


def parse_tairan_ruzhu(rows):
    """泰然立城产业入驻：列与表头一致（购置批次、承租面积等）。"""
    title = rows[0][0] if rows else ""
    headers = rows[1] if len(rows) > 1 else []
    records = []
    for i in range(2, len(rows)):
        r = list(rows[i])
        while len(r) < 28:
            r.append("")
        if not r[2] and not r[3]:
            continue
        records.append(
            {
                "serialNo": r[0],
                "zone": r[1],
                "buildingRoom": r[2],
                "tenantName": r[3],
                "purchaseBatch": r[4],
                "leasedArea": r[5],
                "marketAssessPrice": r[6],
                "rentDiscount": r[7],
                "col8": r[8],
                "contractUnitPrice": r[9],
                "discountUnitPrice": r[10],
                "year1RentAfterDiscount": r[12],
                "year1ContractRent": r[14],
                "year2ContractRent": r[15],
                "year3ContractRent": r[16],
                "year4ContractRent": r[17],
                "year5ContractRent": r[18],
                "startDate": r[19],
                "endDate": r[20],
                "leaseTerm": r[21],
                "rentFreeMonths": r[22],
                "increaseRateYear2": r[23],
                "policyBasis": r[24],
                "rights": r[25],
                "remark": r[26] if len(r) > 26 else "",
            }
        )
    return {"title": title, "headers": headers, "records": records}


def parse_tairan_shangye(rows):
    title = rows[0][0] if rows else ""
    headers = rows[1] if len(rows) > 1 else []
    records = []
    for i in range(2, len(rows)):
        r = rows[i]
        if not any(r[:4]):
            continue
        if r[0] and "合计" in r[0]:
            records.append({"_summaryRow": True, "raw": r})
            continue
        records.append(
            {
                "serialNo": r[0],
                "address": r[1],
                "buildingRoom": r[2],
                "area": r[3],
                "floorOrNo": r[4],
                "status": r[5],
                "progress": r[6],
                "remark": r[7],
            }
        )
    return {"title": title, "headers": headers, "records": records}


def parse_tairan_xiangmu(rows):
    title = rows[0][0] if rows else ""
    records = []
    for i in range(1, len(rows)):
        r = rows[i]
        if not any(r[:4]):
            continue
        records.append({"label": r[0], "value": r[1], "extra": r[2:8]})
    return {"title": title, "records": records}


def main():
    path = find_xlsx()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    out = {
        "sourceFile": os.path.basename(path),
        "sheets": {},
    }
    parsers = [
        parse_baiwang_ruzhu,
        parse_baiwang_shangye,
        parse_baiwang_zujin,
        parse_tairan_ruzhu,
        parse_tairan_shangye,
        parse_tairan_xiangmu,
    ]
    for i, name in enumerate(names):
        if i >= len(parsers):
            break
        ws = wb[name]
        rows = sheet_to_rows(ws)
        out["sheets"][name.strip()] = parsers[i](rows)

    wb.close()

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("OK ->", OUT)
    print("sheets:", list(out["sheets"].keys()))


if __name__ == "__main__":
    main()
